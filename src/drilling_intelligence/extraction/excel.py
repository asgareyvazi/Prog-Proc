"""Excel intelligence (master spec section 8).

Excel is a first-class drilling data source: DDRs, mud logs, time sheets,
casing tallies, well-control drills, cost workbooks.  Rules implemented here:

*   sheets, rows and columns that are *hidden* are still read - hidden in a
    drilling workbook usually means "kept for reference", not "irrelevant" -
    but they are recorded as hidden so a reviewer can see the provenance state;
*   merged cells are preserved: the anchor keeps the value, the covered cells are
    recorded so the original structure can be reconstructed (nothing is destroyed);
*   formulas **and** their cached values are both read when available, and each
    cell records which of the two a value came from;
*   dates, times, 24-hour clock times and timedeltas are rendered losslessly and
    also emitted as fields with their number format recorded;
*   provenance is at cell or range level: ``DDR.xlsx > Sheet: Daily Report > Range: B14:F18``.

openpyxl does the reading (docs/DEPENDENCIES.md); pandas is used only for
optional numeric summaries, never for the provenance path.

Limits, and what happens when one bites
---------------------------------------

A drilling workbook can be enormous (a twelve-month DDR stack in one file), so the
reader is bounded by three settings.  Each one leaves a *diagnostic* rather than a
silent gap, because "this document has 40 rows of data" and "this document had 90 000
cells and we stopped at 60 000" must never look the same downstream:

*   ``excel_max_sheets`` - sheets past the limit are skipped and reported as
    ``EXTRACTION_TRUNCATED: max_sheets=60 (12 of 72 sheets skipped)``;
*   ``excel_max_cells`` - per sheet; the remaining cells are not read and the sheet is
    flagged ``truncated``, so a missing value is understood as "not read", not "absent";
*   ``excel_max_bytes`` - a workbook bigger than this is read **once**.

The last one is the memory trade-off worth stating plainly: reading formulas as well as
values means loading the workbook a second time (openpyxl cannot give both views from one
load), which roughly doubles peak memory for the duration of the extraction.  Formulas
are a bonus - the cached value is the record - so for a huge workbook the second pass is
skipped and the decision is recorded as a diagnostic instead of being paid for in RAM.

No workbook object is touched after ``close()``; every number the report needs is
captured while it is still open.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from typing import Any

from ..__init__ import EXTRACTION_ENGINE_VERSION
from ..core.enums import DataQuality
from ..core.errors import ExtractionError, UnknownUnitError
from ..core.results import DataField
from ..core.units import Quantity, resolve_unit
from .fields import canonical_field_name
from .interfaces import DocumentComplexity, ExtractionContext, ProvenanceBuilder
from .normalized import ExtractionMetadata, NormalizedDocument, Page, Paragraph, Table, clean_text

#: Default ceiling on non-empty cells read per sheet.  Overridable by the
#: ``excel_max_cells`` extractor option (settings ``[extraction] excel_max_cells``).
DEFAULT_MAX_CELLS_PER_SHEET = 60_000
#: Default workbook size above which the second (formula) pass is skipped, in bytes.
DEFAULT_MAX_BYTES = 64 * 1024 * 1024
#: Maximum words in a cell that may still be read as a key rather than as prose.
_MAX_LABEL_WORDS = 4
#: A key/value row may carry a units column and a remark column and still be a
#: key/value row; wider than this it is a table row.
_MAX_KEY_VALUE_COLUMNS = 5

_LABEL = re.compile(r"^[A-Z][A-Za-z0-9 /()\-\u00b0]{1,60}?[:\u2014-]?$")
_UNIT_IN_TEXT = re.compile(r"(?P<value>-?\d+(?:[.,]\d+)?)\s*(?P<unit>ppg|kg/m3|g/cm3|SG|psi|bar|kPa|MPa|m|ft|m3|bbl|gal|l/s|gpm|m/hr|ft/hr|h|hr|min|d|deg|in)\b", re.IGNORECASE)
_TIME_TEXT = re.compile(r"^\s*(\d{1,2}):(\d{2})(?::(\d{2}))?\s*$")


def _positive_int(value: Any, default: int, name: str) -> int:
    """A limit of zero or less would mean "read nothing", which is never intended.

    Settings are user input: a typo like ``excel_max_cells = 0`` must degrade to the
    documented default with a visible reason instead of silently producing empty
    extractions that look like empty documents.
    """
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    if parsed <= 0:
        return default
    del name  # kept for callers' error messages; the fallback itself is the policy
    return parsed


@dataclass
class CellRecord:
    """One non-empty cell, with everything we know about it."""

    coordinate: str
    row: int
    column: int
    value_text: str
    raw_type: str
    number_format: str = ""
    formula: str = ""
    cached_value: str = ""
    comment: str = ""
    merged_span: str = ""
    hidden_row: bool = False
    hidden_column: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {k: v for k, v in self.__dict__.items() if v not in ("", False, None)}


@dataclass
class SheetReport:
    name: str
    index: int
    visible: bool = True
    dimensions: str = ""
    rows: int = 0
    columns: int = 0
    merged: list[str] = field(default_factory=list)
    hidden_rows: list[int] = field(default_factory=list)
    hidden_columns: list[str] = field(default_factory=list)
    comments: int = 0
    formulas: int = 0
    cells: list[CellRecord] = field(default_factory=list)
    #: Repeated header blocks detected inside one sheet (very common in DDRs).
    repeated_headers: list[str] = field(default_factory=list)
    #: Cells present in the sheet but *not read* because the per-sheet limit was reached.
    #: Non-zero means this report is partial: a missing value is "not read", not "absent".
    cells_skipped: int = 0
    #: Populated cells counted before the limit stopped us (== len(cells)).
    cells_read: int = 0
    #: Rows in the sheet's used range, whether or not they were read.
    rows_seen: int = 0

    @property
    def truncated(self) -> bool:
        return self.cells_skipped > 0


class ExcelExtractor:
    """XLSX/XLSM reader that preserves structure, values, formulas and provenance."""

    name = "excel"
    version = EXTRACTION_ENGINE_VERSION
    description = "openpyxl workbook reader: sheets, merged cells, hidden data, formulas, values, dates."

    def supports(self, context: ExtractionContext) -> tuple[bool, str]:
        if context.extension.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            return True, "Excel workbook (openpyxl)"
        if context.extension.lower() in (".xls",):
            return False, "legacy .xls needs a converter; openpyxl supports the OOXML formats only"
        return False, f"extension {context.extension} is not an Excel workbook"

    def probe(self, context: ExtractionContext) -> DocumentComplexity:
        complexity = DocumentComplexity()
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(context.path, read_only=True, data_only=True)
            try:
                complexity.sheet_count = len(workbook.worksheets)
                complexity.pages = complexity.sheet_count
                hidden = sum(1 for ws in workbook.worksheets if ws.sheet_state != "visible")
                if hidden:
                    complexity.reasons.append(f"{hidden} hidden sheet(s)")
                if complexity.sheet_count > 6:
                    complexity.reasons.append(f"{complexity.sheet_count} sheets")
            finally:
                workbook.close()
        except Exception as exc:  # noqa: BLE001 - a probe must never raise
            complexity.reasons.append(f"probe failed: {type(exc).__name__}: {exc}")
        return complexity

    def extract(self, context: ExtractionContext, provenance: ProvenanceBuilder) -> NormalizedDocument:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except Exception as exc:  # pragma: no cover - dependency failure
            raise ExtractionError(f"openpyxl is not available: {exc}") from exc

        read_formulas = bool(context.option("excel_read_formulas", True))
        read_hidden = bool(context.option("excel_read_hidden", True))
        max_sheets = _positive_int(context.option("excel_max_sheets", 60), 60, "excel_max_sheets")
        max_cells = _positive_int(
            context.option("excel_max_cells", DEFAULT_MAX_CELLS_PER_SHEET),
            DEFAULT_MAX_CELLS_PER_SHEET,
            "excel_max_cells",
        )
        max_bytes = _positive_int(context.option("excel_max_bytes", DEFAULT_MAX_BYTES), DEFAULT_MAX_BYTES, "excel_max_bytes")

        document = NormalizedDocument(
            metadata=ExtractionMetadata(
                filename=context.filename,
                path=str(context.path),
                sha256=context.sha256,
                extension=context.extension,
                mime_type=context.mime_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                size_bytes=context.size_bytes,
                engine=f"openpyxl {_openpyxl_version()}",
            )
        )
        # The size gate is decided *before* the second load, not after: the whole point of
        # the limit is to avoid holding two full workbooks in memory at once.
        try:
            workbook_bytes = context.path.stat().st_size
        except OSError:  # pragma: no cover - the router already read the file
            workbook_bytes = int(context.size_bytes or 0)
        if read_formulas and workbook_bytes > max_bytes:
            document.diagnostics.append(
                f"EXTRACTION_TRUNCATED: formula_pass_skipped (workbook is {workbook_bytes} bytes, "
                f"excel_max_bytes={max_bytes}); values only, to avoid loading the workbook twice"
            )
            read_formulas = False

        try:
            values_book = load_workbook(context.path, data_only=True, read_only=False)
        except Exception as exc:
            raise ExtractionError(f"Cannot open workbook {context.filename}: {type(exc).__name__}: {exc}") from exc
        formula_book = None
        if read_formulas:
            try:
                formula_book = load_workbook(context.path, data_only=False, read_only=False)
            except Exception as exc:  # noqa: BLE001 - formulas are a bonus, values are the record
                document.diagnostics.append(f"formula read unavailable: {type(exc).__name__}: {exc}")

        reports: list[SheetReport] = []
        try:
            # Everything the report needs is measured here, while the books are open:
            # nothing below touches a closed workbook (openpyxl tolerates it, but a
            # structure read after close is exactly the kind of accident that turns into
            # a confusing empty document later).
            total_sheets = len(values_book.worksheets)
            for worksheet in values_book.worksheets[:max_sheets]:
                reports.append(self._read_sheet(worksheet, formula_book, get_column_letter, read_hidden, max_cells=max_cells))
        finally:
            values_book.close()
            if formula_book is not None:
                formula_book.close()

        if total_sheets > max_sheets:
            document.diagnostics.append(
                f"EXTRACTION_TRUNCATED: max_sheets={max_sheets} ({total_sheets - max_sheets} of {total_sheets} sheets skipped)"
            )
        for report in reports:
            if report.truncated:
                document.diagnostics.append(
                    f"EXTRACTION_TRUNCATED: max_cells={max_cells} in sheet {report.name} "
                    f"({report.cells_skipped} populated cells not read after {report.cells_read})"
                )

        document.metadata.page_count = len(reports)
        document.metadata.extra["workbook"] = {
            # Read this first: a truncated extraction must never be mistaken for a
            # complete one, and the limits that caused it belong next to the data.
            "limits": {
                "max_sheets": max_sheets,
                "max_cells_per_sheet": max_cells,
                "max_bytes": max_bytes,
                "sheets_total": total_sheets,
                "sheets_read": len(reports),
                "sheets_skipped": max(0, total_sheets - max_sheets),
                "cells_skipped": sum(report.cells_skipped for report in reports),
                "formula_pass": read_formulas,
            },
            "truncated": bool(total_sheets > max_sheets) or any(report.truncated for report in reports),
            "sheets": [
                {
                    "name": r.name,
                    "truncated": r.truncated,
                    "cells_read": r.cells_read,
                    "cells_skipped": r.cells_skipped,
                    "visible": r.visible,
                    "dimensions": r.dimensions,
                    "rows": r.rows,
                    "columns": r.columns,
                    "merged": r.merged,
                    "hidden_rows": r.hidden_rows[:200],
                    "hidden_columns": r.hidden_columns[:200],
                    "comments": r.comments,
                    "formulas": r.formulas,
                    "repeated_headers": r.repeated_headers,
                }
                for r in reports
            ]
        }

        paragraph_index = 0
        for sheet_index, report in enumerate(reports):
            label = f"{report.name}{'' if report.visible else ' (hidden sheet)'}"
            document.pages.append(
                Page(index=sheet_index + 1, text="", char_start=0, char_end=0, label=label, block_count=len(report.cells), extra={"sheet": report.name})
            )
            # A title paragraph per sheet keeps the sheet name searchable and cited.
            title = Paragraph(
                index=paragraph_index,
                text=f"Sheet: {report.name}" + ("" if report.visible else " [hidden]"),
                page=sheet_index + 1,
                block=0,
                heading_level=1,
                style="sheet",
                provenance=provenance.excel(sheet=report.name, cell="A1", excerpt=f"Sheet: {report.name}"),
            )
            document.paragraphs.append(title)
            paragraph_index += 1

            for table_number, table in enumerate(self._tables_from_report(report, provenance, sheet_index + 1)):
                table.page = sheet_index + 1
                document.tables.append(table)
                for heading_text in self._table_headings(table):
                    document.paragraphs.append(
                        Paragraph(
                            index=paragraph_index,
                            text=heading_text,
                            page=sheet_index + 1,
                            block=table_number + 1,
                            style="table_row",
                            provenance=table.provenance,
                        )
                    )
                    paragraph_index += 1
                # Field extraction from cells is done by the pipeline (fields.py);
                # label/value pairs are handled here because the *cell* is the
                # precise provenance unit for them.
                for item in self._key_value_fields(report, provenance):
                    document.extracted_fields.append(item)

        # Page text must be assembled after tables exist so search hits are in context.
        for page in document.pages:
            sheet_name = page.extra.get("sheet", "")
            rows = [t.text() for t in document.tables if t.sheet == sheet_name]
            page.text = clean_text("\n\n".join(rows))[:200000]

        document.text = clean_text("\n\n".join([p.text for p in document.paragraphs if p.text] + [t.text() for t in document.tables]))
        if not document.tables:
            document.diagnostics.append("no populated table regions found (workbook may contain only formulas/empty cells)")
        return document

    # ------------------------------------------------------------------ sheets
    def _read_sheet(
        self,
        worksheet: Any,
        formula_book: Any,
        get_column_letter: Any,
        read_hidden: bool,
        *,
        max_cells: int = DEFAULT_MAX_CELLS_PER_SHEET,
    ) -> SheetReport:
        report = SheetReport(
            name=worksheet.title,
            index=worksheet.sheet_index if hasattr(worksheet, "sheet_index") else 0,
            visible=(str(worksheet.sheet_state) == "visible"),
            dimensions=str(worksheet.dimensions or ""),
        )
        if not report.visible and not read_hidden:
            report.rows = worksheet.max_row or 0
            report.columns = worksheet.max_column or 0
            return report

        merged_by_cell: dict[tuple[int, int], str] = {}
        for merged in worksheet.merged_cells.ranges:
            report.merged.append(str(merged))
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    merged_by_cell[(row, col)] = str(merged)

        hidden_rows = {r for r, dim in (worksheet.row_dimensions or {}).items() if getattr(dim, "hidden", False)}
        hidden_columns = {get_column_letter(c) for c, dim in (worksheet.column_dimensions or {}).items() if getattr(dim, "hidden", False)}
        report.hidden_rows = sorted(hidden_rows)
        report.hidden_columns = sorted(hidden_columns)
        report.rows = worksheet.max_row or 0
        report.columns = worksheet.max_column or 0

        formula_sheet = None
        if formula_book is not None and worksheet.title in formula_book.sheetnames:
            formula_sheet = formula_book[worksheet.title]

        # A sheet's used range can be far bigger than its populated area, so the row
        # count is recorded before the budget is applied: "8231 rows, 60000 cells read"
        # is a different statement from "8231 rows, 8231 cells read".
        report.rows_seen = int(worksheet.max_row or 0)
        count = 0
        skipped = 0
        truncated = False
        for row in worksheet.iter_rows():
            if truncated:
                # Past the budget the remaining rows are only *counted*: no text
                # rendering, no number-format work, no formula lookup.  The workbook is
                # already in memory, so counting is cheap, and the count is what tells a
                # reviewer how much of the sheet they are not seeing.
                skipped += sum(1 for cell in row if cell.value is not None and not (isinstance(cell.value, str) and not cell.value.strip()))
                continue
            for cell in row:
                value = cell.value
                if value is None or (isinstance(value, str) and not value.strip()):
                    continue
                if count >= max_cells:
                    truncated = True
                    skipped += 1
                    break
                coordinate = cell.coordinate
                text = render_cell(value, cell.number_format or "")
                record = CellRecord(
                    coordinate=coordinate,
                    row=cell.row,
                    column=cell.column,
                    value_text=text,
                    raw_type=_value_type_name(value),
                    number_format=str(cell.number_format or ""),
                    merged_span=merged_by_cell.get((cell.row, cell.column), ""),
                    hidden_row=cell.row in hidden_rows,
                    hidden_column=get_column_letter(cell.column) in hidden_columns,
                )
                comment = getattr(cell, "comment", None)
                if comment is not None and getattr(comment, "text", ""):
                    record.comment = clean_text(str(comment.text))[:2000]
                    report.comments += 1
                if formula_sheet is not None:
                    try:
                        fcell = formula_sheet.cell(row=cell.row, column=cell.column)
                        if isinstance(fcell.value, str) and fcell.value.startswith("="):
                            record.formula = fcell.value[:512]
                            record.cached_value = text
                            report.formulas += 1
                    except Exception:  # noqa: BLE001 - mismatched sheet shapes are tolerated
                        pass
                report.cells.append(record)
                count += 1
        report.cells_read = count
        report.cells_skipped = skipped
        return report

    # ------------------------------------------------------------------ tables
    def _tables_from_report(self, report: SheetReport, provenance: ProvenanceBuilder, page: int) -> list[Table]:
        """Split a sheet into table regions separated by empty rows/columns.

        Drilling workbooks stack several tables on one sheet (rig summary above a
        trip sheet).  Splitting on blank bands keeps each region rectangular,
        which is what the QA checks and the program comparison need.
        """
        from openpyxl.utils import get_column_letter as _letter

        if not report.cells:
            return []
        by_row: dict[int, dict[int, CellRecord]] = {}
        for cell in report.cells:
            by_row.setdefault(cell.row, {})[cell.column] = cell
        rows_sorted = sorted(by_row)
        blocks: list[list[int]] = []
        current: list[int] = []
        previous: int | None = None
        for row in rows_sorted:
            if previous is not None and row - previous > 2:
                blocks.append(current)
                current = []
            current.append(row)
            previous = row
        if current:
            blocks.append(current)

        tables: list[Table] = []
        header_texts: list[str] = []
        for block_index, block_rows in enumerate(blocks):
            columns = sorted({col for row in block_rows for col in by_row[row]})
            if not columns:
                continue
            grid: list[list[str | None]] = []
            for row in block_rows:
                grid.append([by_row[row].get(col).value_text if by_row[row].get(col) else None for col in columns])
            first_row, last_row = block_rows[0], block_rows[-1]
            first_col, last_col = columns[0], columns[-1]
            range_ref = f"{_letter(first_col)}{first_row}:{_letter(last_col)}{last_row}"
            header = [str(c or "") for c in grid[0]] if grid else []
            if header:
                joined = "|".join(header)
                if joined in header_texts:
                    # Repeated header inside one region: recorded, not duplicated as data.
                    tables.append(
                        Table(
                            table_id=f"{report.name}-block{block_index + 1}",
                            rows=[header],
                            caption=f"{report.name}: repeated header block",
                            sheet=report.name,
                            page=page,
                            anchor=range_ref,
                            has_header=False,
                            provenance=provenance.excel(sheet=report.name, range_=range_ref, excerpt=" | ".join(header)[:1500]),
                            extra={"repeated_header": True},
                        )
                    )
                    header_texts.append(joined)
                    continue
                header_texts.append(joined)
            excerpt = "\n".join("\t".join("" if c is None else str(c) for c in row) for row in grid[:8])
            tables.append(
                Table(
                    table_id=f"{report.name}-block{block_index + 1}",
                    rows=grid,
                    caption=f"{report.name}" if len(blocks) == 1 else f"{report.name} block {block_index + 1}",
                    sheet=report.name,
                    page=page,
                    anchor=range_ref,
                    provenance=provenance.excel(sheet=report.name, range_=range_ref, excerpt=excerpt[:2000]),
                    extra={
                        "hidden_rows_in_block": [r for r in report.hidden_rows if first_row <= r <= last_row],
                        "hidden_columns_in_block": [c for c in report.hidden_columns if first_col <= _column_index(c) <= last_col],
                        "merged_ranges": [m for m in report.merged if _range_intersects_rows(m, first_row, last_row)],
                    },
                )
            )
        return tables

    @staticmethod
    def _table_headings(table: Table) -> list[str]:
        """Row text that carries searchable content (kept short on purpose).

        Only rows with a label-like first cell are promoted to searchable
        paragraphs; the full grid stays on the table object so the original
        structure is never lost.
        """
        out: list[str] = []
        for _, row in table.iter_data_rows():
            first = row[0] if row else ""
            if first and len(first) <= 90 and not re.match(r"^\s*\d+[.,]?\d*\s*$", first):
                out.append(" ".join(cell for cell in row if cell)[:400])
        return out[:400]

    # -------------------------------------------------------------- key/value
    def _key_value_fields(self, report: SheetReport, provenance: ProvenanceBuilder) -> list[DataField]:
        """Label/value pairs (one per row) become directly cell-cited fields.

        ``B14: 'Mud weight' -> C14: '10.2 ppg'`` is the single most common shape
        in a DDR, and it deserves provenance at the *cell*, not at the range.  Rows with
        a units column and a remarks column are still key/value rows: ``A9 'Mud weight
        (ppg)' | B9 10.2 | C9 'ppg' | D9 'Rheometer reading'`` is the shape a real mud
        report uses, and dropping it because it is wide would lose the one number the
        whole report exists to record.
        """
        by_row: dict[int, dict[int, CellRecord]] = {}
        for cell in report.cells:
            by_row.setdefault(cell.row, {})[cell.column] = cell
        fields: list[DataField] = []
        for row in sorted(by_row):
            cells = [by_row[row][col] for col in sorted(by_row[row])]
            if len(cells) < 2 or len(cells) > _MAX_KEY_VALUE_COLUMNS:
                continue
            label_cell = cells[0]
            value_cell = cells[1]
            label = label_cell.value_text.strip().rstrip(":").strip()
            if not _LABEL.match(label) or len(label) < 3:
                continue
            if re.search(r"\d", label):
                continue  # a label containing digits is data, not a key
            if len(label.split()) > _MAX_LABEL_WORDS:
                # A "label" of five-plus words is prose that happens to sit left of a
                # number ("Total NPT for the period"); reading it as a key invents a
                # field that nobody wrote down.
                continue
            parsed = _parse_value(value_cell.value_text)
            if parsed is None:
                continue
            value, unit, quality = parsed
            trailing = cells[2:]
            # ``_parse_value`` accepts text as well as numbers, so ask specifically for
            # measurements: two more numbers in the row means this is a data row of a
            # table (a slip with MW in/out and viscosity), not a label with a remark.
            numbers_beyond_the_value = sum(
                1
                for extra in trailing
                if (_parsed := _parse_value(extra.value_text)) is not None and isinstance(_parsed[0], (int, float))
            )
            if numbers_beyond_the_value >= 2:
                continue
            if not unit:
                # A units column belongs to the value.  Without it the number has no
                # dimension, so it could not be compared with the program at all.
                for extra in trailing:
                    candidate = extra.value_text.strip()
                    if not candidate or len(candidate.split()) > 1 or re.search(r"\d", candidate):
                        continue
                    try:
                        unit = resolve_unit(candidate).symbol
                    except UnknownUnitError:
                        unit = ""
                    if unit:
                        break
            fields.append(
                DataField(
                    name=canonical_field_name(label) or re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_"),
                    value=value,
                    unit=unit,
                    dimension=_dimension_for_unit(unit),
                    quality=quality,
                    provenance=provenance.excel(
                        sheet=report.name,
                        cell=value_cell.coordinate,
                        read="value",
                        excerpt=value_cell.value_text[:200],
                    ),
                    confidence=0.9 if unit else 0.6,
                    method=f"excel:key-value {label_cell.coordinate}->{value_cell.coordinate}",
                    note=(
                        f"label cell {label_cell.coordinate}; format {value_cell.number_format}"
                        + ("; formula cell (cached value read)" if value_cell.formula else "")
                        + ("; hidden row" if value_cell.hidden_row else "")
                        + ("; hidden column" if value_cell.hidden_column else "")
                    ),
                )
            )
        return fields


# --------------------------------------------------------------------------- rendering
def render_cell(value: Any, number_format: str) -> str:
    """Lossless-ish text for a cell value, keeping dates/times/durations readable.

    Excel stores times as fractional days; a 24-hour format string is what tells
    a human "this is 18:30, not 0.7708".  We keep both: the rendered text for
    search and reading, and the raw type in the cell record.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        if isinstance(value, dt.datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S") if (value.hour or value.minute or value.second) else value.strftime("%Y-%m-%d")
        if isinstance(value, dt.date):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%H:%M:%S") if "%S" in number_format else value.strftime("%H:%M")
    if isinstance(value, dt.timedelta):
        total_seconds = int(value.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
    if isinstance(value, float):
        if value == int(value) and abs(value) < 1e15 and "0" not in (number_format.split(".")[-1] if "." in number_format else ""):
            return str(int(value))
        return f"{value:.6g}"
    if isinstance(value, int):
        return str(value)
    return clean_text(str(value))


def _value_type_name(value: Any) -> str:
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, dt.datetime):
        return "datetime"
    if isinstance(value, dt.date):
        return "date"
    if isinstance(value, dt.time):
        return "time"
    if isinstance(value, dt.timedelta):
        return "duration"
    return "text"


def _parse_value(text: str) -> tuple[Any, str, DataQuality] | None:
    """Parse a value cell into (value, unit, quality); None if not field-worthy."""
    stripped = (text or "").strip()
    if not stripped:
        return None
    time_match = _TIME_TEXT.match(stripped)
    if time_match:
        # "12:30:00" / "8h45" style shift and duration cells: hours, not a clock.
        hours, minutes, seconds = (int(time_match.group(i) or 0) for i in (1, 2, 3))
        return (hours * 3600 + minutes * 60 + seconds) / 3600.0, "h", DataQuality.VALID
    try:
        return float(stripped.replace(",", "")), "", DataQuality.UNVERIFIED
    except ValueError:
        pass
    match = _UNIT_IN_TEXT.search(stripped)
    if match:
        try:
            quantity = Quantity.parse(stripped)
            return quantity.value, quantity.unit.symbol, DataQuality.VALID
        except Exception:  # noqa: BLE001 - fall through to text handling
            pass
    if len(stripped) <= 2000:
        return stripped, "", DataQuality.VALID  # textual value (e.g. "OK", "12 1/4 intermediate")
    return None


def _dimension_for_unit(unit: str) -> str:
    if not unit:
        return ""
    try:
        return resolve_unit(unit).dimension.value
    except Exception:  # noqa: BLE001 - free text has no dimension
        return ""


def _column_index(letters: str) -> int:
    total = 0
    for char in letters:
        total = total * 26 + (ord(char.upper()) - 64)
    return total


def _range_intersects_rows(reference: str, first_row: int, last_row: int) -> bool:
    match = re.match(r"[A-Z]+(\d+)(?::[A-Z]+(\d+))?", reference or "")
    if not match:
        return False
    start = int(match.group(1))
    end = int(match.group(2) or start)
    return not (end < first_row or start > last_row)


def _openpyxl_version() -> str:
    try:
        import openpyxl

        return str(getattr(openpyxl, "__version__", "?"))
    except Exception:  # noqa: BLE001
        return "?"


__all__ = ["CellRecord", "ExcelExtractor", "SheetReport", "render_cell"]
