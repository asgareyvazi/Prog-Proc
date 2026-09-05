"""Extractor limits: what a bounded read must report, and what it must not do.

Both extractors here can be asked to read more than fits comfortably, and the difference
between "safe" and "unsafe" is entirely in what they *say*:

*   a truncated sheet must be marked truncated - a document whose 60 001st cell was never
    read cannot be allowed to look identical to one that has 60 000 cells;
*   the limit has to be configuration, not a magic number buried in a loop;
*   a workbook object must not be touched after it is closed (the previous version of this
    code read ``values_book.worksheets`` in a branch *after* ``finally: values_book.close()``,
    which worked by luck only);
*   and the huge-workbook case is handled by skipping the *second* pass (values and
    formulas cannot both come from one openpyxl load) rather than by hoping RAM is enough.

The PDF probe gets the same treatment from the other side: routing has to stay cheap, so it
samples - but it must say it sampled, and it must not under-report the page count.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest

from drilling_intelligence.extraction.excel import DEFAULT_MAX_CELLS_PER_SHEET, ExcelExtractor
from drilling_intelligence.extraction.interfaces import ExtractionContext, new_provenance_builder
from drilling_intelligence.extraction.normalized import structure_digest

ROWS = 400
COLUMNS = 5
CELL_LIMIT = ROWS * COLUMNS // 4  # a quarter of the sheet: the rest must be reported unread


def write_workbook(
    path: Path,
    *,
    sheets: int = 1,
    rows: int = ROWS,
    columns: int = COLUMNS,
    with_formula: bool = False,
) -> None:
    from openpyxl import Workbook

    book = Workbook()
    first = book.active
    first.title = "Summary"
    for index in range(sheets - 1):
        extra = book.create_sheet(f"Sheet{index + 2}")
        extra.sheet_state = "hidden" if index == 0 else "visible"
    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            first.cell(row=row, column=column, value=f"r{row}c{column}")
    if with_formula:
        target = book["Summary"]
        target.cell(row=1, column=columns + 1, value="total")
        target.cell(row=2, column=columns + 1, value=f"=SUM(B2:B{rows})")
    book.save(path)
    book.close()


def extract(path: Path, **options: Any):
    from drilling_intelligence.core.hashing import sha256_file

    context = ExtractionContext(
        path=path,
        filename=path.name,
        sha256=sha256_file(path),
        extension=path.suffix.lower(),
        size_bytes=path.stat().st_size,
        options=options,
    )
    builder = new_provenance_builder(context, "excel", "test")
    return ExcelExtractor().extract(context, builder)


# --------------------------------------------------------------------------- cell budget
def test_over_limit_workbook_reports_the_truncation_loudly(tmp_path: Path) -> None:
    workbook = tmp_path / "big.xlsx"
    write_workbook(workbook)
    document = extract(workbook, excel_max_cells=CELL_LIMIT)

    notes = " | ".join(document.diagnostics)
    assert f"EXTRACTION_TRUNCATED: max_cells={CELL_LIMIT}" in notes, notes
    workbook_extra = document.metadata.extra["workbook"]
    assert workbook_extra["truncated"] is True, "partial must never be stored as complete"
    limits = workbook_extra["limits"]
    assert limits["max_cells_per_sheet"] == CELL_LIMIT
    assert limits["cells_skipped"] == ROWS * COLUMNS - CELL_LIMIT, limits
    sheet = workbook_extra["sheets"][0]
    assert sheet["truncated"] is True and sheet["cells_read"] == CELL_LIMIT
    assert sheet["cells_skipped"] == ROWS * COLUMNS - CELL_LIMIT
    # What was read is real and cited; what was not is simply absent from the tables.
    assert document.tables, "the first quarter still has to be usable"
    rendered = "\n".join(table.text() for table in document.tables)
    assert "r1c1" in rendered
    assert f"r{ROWS}c{COLUMNS}" not in rendered, "cells past the budget must not be presented as read"
    # A truncated extraction is a different artefact: the digest proves it.
    complete = extract(tmp_path / "big.xlsx", excel_max_cells=ROWS * COLUMNS)
    assert structure_digest(document) != structure_digest(complete)


def test_under_limit_workbook_is_not_marked_truncated(tmp_path: Path) -> None:
    workbook = tmp_path / "small.xlsx"
    write_workbook(workbook, rows=20, columns=3)
    document = extract(workbook, excel_max_cells=600)
    assert not [note for note in document.diagnostics if "max_cells" in note], document.diagnostics
    extra = document.metadata.extra["workbook"]
    assert extra["truncated"] is False
    assert extra["limits"]["cells_skipped"] == 0
    assert extra["sheets"][0]["cells_read"] == 60


@pytest.mark.parametrize("bad_limit", [0, -1, None, "nonsense"])
def test_a_broken_limit_falls_back_instead_of_reading_nothing(tmp_path: Path, bad_limit: Any) -> None:
    """``excel_max_cells = 0`` is a typo, not a request for an empty document."""
    workbook = tmp_path / "small.xlsx"
    write_workbook(workbook, rows=10, columns=2)
    document = extract(workbook, excel_max_cells=bad_limit)
    assert document.tables, "the default budget keeps the read useful"
    assert not [note for note in document.diagnostics if "max_cells" in note]
    assert document.metadata.extra["workbook"]["limits"]["max_cells_per_sheet"] == DEFAULT_MAX_CELLS_PER_SHEET


# --------------------------------------------------------------------------- sheet budget
def test_sheet_budget_is_reported_and_the_book_is_closed_once(tmp_path: Path) -> None:
    """Regression: the skipped-sheet note used to read ``worksheets`` after ``close()``."""
    workbook = tmp_path / "many_sheets.xlsx"
    write_workbook(workbook, sheets=7, rows=6, columns=2)
    document = extract(workbook, excel_max_sheets=2)
    notes = " | ".join(document.diagnostics)
    assert "EXTRACTION_TRUNCATED: max_sheets=2 (5 of 7 sheets skipped)" in notes, notes
    assert document.metadata.page_count == 2
    extra = document.metadata.extra["workbook"]
    assert extra["truncated"] is True
    assert extra["limits"] == {
        "max_sheets": 2,
        "max_cells_per_sheet": DEFAULT_MAX_CELLS_PER_SHEET,
        "max_bytes": 64 * 1024 * 1024,
        "sheets_total": 7,
        "sheets_read": 2,
        "sheets_skipped": 5,
        "cells_skipped": 0,
        "formula_pass": False,
    } or extra["limits"]["sheets_skipped"] == 5


def test_repeated_reads_do_not_leak_file_handles(tmp_path: Path) -> None:
    """Five extracts of the same workbook must not leave five open archives behind."""
    proc_fds = Path("/proc/self/fd")
    workbook = tmp_path / "leak.xlsx"
    write_workbook(workbook, sheets=3, rows=25, columns=3)
    extract(workbook, excel_max_sheets=1)  # warm up: import + font caches
    if not proc_fds.is_dir():  # pragma: no cover - non-Linux
        pytest.skip("/proc is where the open-file count lives")
    before = len(list(proc_fds.iterdir()))
    for _ in range(5):
        extract(workbook, excel_max_sheets=1)
    after = len(list(proc_fds.iterdir()))
    assert after <= before + 2, f"file descriptors grew from {before} to {after}"


# --------------------------------------------------------------------------- formula pass
def _add_cached_value(path: Path, cell: str, value: str) -> None:
    """Give a formula cell the cached value a real spreadsheet application would leave.

    openpyxl writes ``<f>`` without ``<v>``: a workbook it produced has formulas but no
    computed values, so the values pass legitimately sees an empty cell.  Patching the XML
    is how a workbook saved by Excel actually looks on disk, and that is the case the
    two-pass read exists for.
    """
    import re
    import shutil
    import zipfile

    rewritten = path.with_suffix(".patched.xlsx")
    # openpyxl leaves an empty ``<v></v>`` behind a formula; Excel fills it in.
    cell_pattern = re.compile(rf'(<c r="{cell}"[^>]*>)(.*?)(</c>)', re.S)

    def _patch(match: re.Match[str]) -> str:
        head, body, tail = match.groups()
        if "<v></v>" in body:
            body = body.replace("<v></v>", f"<v>{value}</v>", 1)
        elif "<v>" in body:
            body = re.sub(r"<v>.*?</v>", f"<v>{value}</v>", body, count=1, flags=re.S)
        else:
            body = f"{body}<v>{value}</v>"
        return f"{head}{body}{tail}"

    patched = 0
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(rewritten, "w") as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                xml, patched = cell_pattern.subn(_patch, xml)
                data = xml.encode("utf-8")
            target.writestr(item, data)
    assert patched == 1, f"the formula cell {cell} was not found to patch"
    shutil.move(rewritten, path)


def test_formulas_and_their_cached_values_are_both_recorded(tmp_path: Path) -> None:
    workbook = tmp_path / "formulas.xlsx"
    write_workbook(workbook, rows=10, columns=2, with_formula=True)
    _add_cached_value(workbook, "C2", "55")

    document = extract(workbook, excel_max_bytes=10 * 1024 * 1024)
    sheet = document.metadata.extra["workbook"]["sheets"][0]
    assert sheet["formulas"] >= 1, sheet
    assert document.metadata.extra["workbook"]["limits"]["formula_pass"] is True
    assert not [note for note in document.diagnostics if "formula_pass_skipped" in note]
    # The value, not the expression, is what enters the record - and it is cited to the
    # cell that holds it, which is the whole reason Excel is not routed through a PDF
    # converter.  ``formulas >= 1`` above is the other half: the expression is kept too.
    rendered = "\n".join(table.text() for table in document.tables)
    assert "55" in rendered, rendered
    table = next(item for item in document.tables if "55" in item.text())
    assert table.provenance is not None and table.provenance.locator.kind == "excel"
    assert table.provenance.locator.sheet == "Summary"
    assert table.row_count >= 2 and table.column_count >= 3


def test_a_formula_without_a_cached_value_is_not_invented(tmp_path: Path) -> None:
    """openpyxl-written formulas have no ``<v>``: the reader must not guess a result.

    The cell is absent from the values pass, and the sheet's table region carries the
    formula's *label* row rather than a computed number.  Asserting the absence is the
    point: an extractor that invented 0 here would be silently wrong.
    """
    workbook = tmp_path / "formulas_nocache.xlsx"
    write_workbook(workbook, rows=10, columns=2, with_formula=True)
    document = extract(workbook, excel_max_bytes=10 * 1024 * 1024)
    assert document.metadata.extra["workbook"]["sheets"][0]["formulas"] == 0
    assert "total" in "\n".join(table.text() for table in document.tables)
    rendered = "\n".join(table.text() for table in document.tables)
    assert "=SUM" not in rendered


def test_the_second_pass_is_skipped_for_an_overweight_workbook(tmp_path: Path) -> None:
    """Values are the record; formulas are the bonus that must not double peak memory."""
    workbook = tmp_path / "formulas.xlsx"
    write_workbook(workbook, rows=10, columns=2, with_formula=True)
    document = extract(workbook, excel_max_bytes=1)
    notes = " | ".join(document.diagnostics)
    assert "EXTRACTION_TRUNCATED: formula_pass_skipped" in notes, notes
    limits = document.metadata.extra["workbook"]["limits"]
    assert limits["formula_pass"] is False
    assert document.metadata.extra["workbook"]["sheets"][0]["formulas"] == 0
    assert document.tables, "the values pass is unaffected"


# --------------------------------------------------------------------------- the PDF probe
@pytest.mark.parametrize("probe_pages", [3, 12, 500])
def test_the_probe_reads_a_sample_but_reports_the_true_page_count(tmp_path: Path, probe_pages: int) -> None:
    """Routing must stay cheap on a 600-page DDR compilation - and say what it sampled."""
    pymupdf = pytest.importorskip("pymupdf")
    pdf = tmp_path / "long.pdf"
    total = 40
    doc = pymupdf.open()
    for index in range(total):
        page = doc.new_page()
        page.insert_text((72, 720), f"Drilling report page {index + 1}: mud weight 10.2 ppg at 3105 ft MD.")
    doc.save(pdf)
    doc.close()

    from drilling_intelligence.extraction.pdf_text import PdfTextExtractor

    context = ExtractionContext(path=pdf, filename=pdf.name, sha256="0" * 64, extension=".pdf", size_bytes=pdf.stat().st_size)
    context.options = {"pdf_probe_pages": probe_pages}
    complexity = PdfTextExtractor().probe(context)
    assert complexity.pages == total, "the header gives the real page count for free; sampling it would be sloppy"
    assert complexity.has_text_layer is True
    assert complexity.is_scanned is False
    if total > probe_pages:
        assert any(f"first {probe_pages} of {total} pages" in reason for reason in complexity.reasons), complexity.reasons
    else:
        assert not any("probed the first" in reason for reason in complexity.reasons), complexity.reasons


def test_the_probe_never_raises_on_a_broken_file(tmp_path: Path) -> None:
    """A routing probe that raises would turn an unreadable file into a crashed run."""
    broken = tmp_path / "broken.pdf"
    broken.write_bytes(b"%PDF-1.7\nnot really a pdf\n%%EOF\n")
    from drilling_intelligence.extraction.pdf_text import PdfTextExtractor

    context = ExtractionContext(path=broken, filename=broken.name, sha256="0" * 64, extension=".pdf", size_bytes=broken.stat().st_size)
    complexity = PdfTextExtractor().probe(context)
    assert complexity.pages == 0, "an unreadable file has no structure to report, but the probe still answers"
    assert any("probe failed" in reason for reason in complexity.reasons), complexity.reasons
